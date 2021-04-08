import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook('Financial Data.xlsx')

sheet = wb['2020']
cell = sheet['e2']

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 5)

values = Reference(sheet,
                   min_row=2,
                   max_row=sheet.max_row,
                   min_col=5,
                   max_col=5)

chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'c704')